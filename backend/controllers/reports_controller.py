from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt
from extensions import db
from datetime import datetime
import io, openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

# ─── Palette ────────────────────────────────────────────────────
C_HEADER  = '1A2341'   # dark navy — header row
C_TITLE   = '1E40AF'   # blue — section titles
C_TOTAL   = 'DBEAFE'   # light blue — totals row
C_ODD     = 'F8FAFC'   # light grey — alternating rows
C_GREEN   = '059669'
C_RED     = 'DC2626'
C_AMBER   = 'D97706'

def _fill(hex_):
    return PatternFill('solid', start_color=hex_, end_color=hex_)

def _border():
    thin = Side(style='thin', color='E2E8F0')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _hdr_font(white=True):
    return Font(name='Arial', bold=True, size=10, color='FFFFFF' if white else C_TITLE)

def _body_font(bold=False, color='111827'):
    return Font(name='Arial', bold=bold, size=10, color=color)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _right():
    return Alignment(horizontal='right', vertical='center')

FMT_VND  = '#,##0'
FMT_PCT  = '0.0%'

def _write_header(ws, title, date_from, date_to, cols):
    """Write report title block and column headers, return next row."""
    # Row 1 – report title
    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    c = ws['A1']
    c.value = title
    c.font  = Font(name='Arial', bold=True, size=14, color=C_TITLE)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Row 2 – date range
    ws.merge_cells(f'A2:{get_column_letter(len(cols))}2')
    c = ws['A2']
    c.value = f'Kỳ báo cáo: {date_from}  →  {date_to}     Xuất lúc: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    c.font  = Font(name='Arial', italic=True, size=9, color='6B7280')
    c.alignment = _center()
    ws.row_dimensions[2].height = 16

    # Row 3 – blank
    ws.row_dimensions[3].height = 6

    # Row 4 – column headers
    for ci, h in enumerate(cols, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font      = _hdr_font()
        cell.fill      = _fill(C_HEADER)
        cell.alignment = _center() if ci > 1 else Alignment(horizontal='left', vertical='center')
        cell.border    = _border()
    ws.row_dimensions[4].height = 22
    return 5   # data starts at row 5

def _write_rows(ws, start_row, rows_data, col_fmts=None):
    """Write data rows with alternating fill. col_fmts: list of (fmt, color) per col."""
    for ri, row in enumerate(rows_data):
        fill = _fill(C_ODD) if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=start_row + ri, column=ci, value=val)
            cell.border = _border()
            if fill: cell.fill = fill
            if col_fmts and ci <= len(col_fmts):
                fmt, color = col_fmts[ci - 1]
                if fmt:   cell.number_format = fmt
                if color: cell.font = _body_font(color=color)
                else:     cell.font = _body_font()
            else:
                cell.font = _body_font()
            if ci > 1: cell.alignment = _right()
            else:       cell.alignment = Alignment(horizontal='left', vertical='center')

def _write_total(ws, row_idx, cells):
    """Write a styled total row. cells = list of values for each column."""
    for ci, val in enumerate(cells, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.fill   = _fill(C_TOTAL)
        cell.font   = Font(name='Arial', bold=True, size=10, color=C_TITLE)
        cell.border = _border()
        cell.alignment = _right() if ci > 1 else Alignment(horizontal='left', vertical='center')

def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _autofreeze(ws, row=5, col=2):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ─── Helpers: admin check ───────────────────────────────────────
def _admin():
    claims = get_jwt()
    return claims.get('role') == 'admin'

def _params():
    tab       = request.args.get('tab', 'revenue')
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    group_by  = request.args.get('group_by', 'day')
    return tab, date_from, date_to, group_by

# ─── Date label format ──────────────────────────────────────────
def _date_fmt(group_by):
    return "DATE_FORMAT(created_at, '%d/%m/%Y')" if group_by == 'day' else "DATE_FORMAT(created_at, '%m/%Y')"

def _group_expr(group_by):
    return "DATE(created_at)" if group_by == 'day' else "DATE_FORMAT(created_at,'%Y-%m')"


# ════════════════════════════════════════════════════════════════
# GET /admin/reports/data  — fetch report data
# ════════════════════════════════════════════════════════════════
@reports_bp.route('/data', methods=['GET'])
@jwt_required()
def get_report_data():
    if not _admin():
        return jsonify({'success': False, 'message': 'Không có quyền'}), 403

    tab, date_from, date_to, group_by = _params()
    if not date_from or not date_to:
        return jsonify({'success': False, 'message': 'Thiếu khoảng thời gian'}), 400

    try:
        data = _build_data(tab, date_from, date_to, group_by)
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ════════════════════════════════════════════════════════════════
# GET /admin/reports/export — export Excel
# ════════════════════════════════════════════════════════════════
@reports_bp.route('/export', methods=['GET'])
@jwt_required()
def export_excel():
    if not _admin():
        return jsonify({'success': False, 'message': 'Không có quyền'}), 403

    tab, date_from, date_to, group_by = _params()
    if not date_from or not date_to:
        return jsonify({'success': False, 'message': 'Thiếu khoảng thời gian'}), 400

    try:
        data = _build_data(tab, date_from, date_to, group_by)
        wb   = _build_excel(tab, date_from, date_to, group_by, data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f'bao_cao_{tab}_{date_from}_{date_to}.xlsx'
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname,
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ════════════════════════════════════════════════════════════════
# Data builders
# ════════════════════════════════════════════════════════════════
def _build_data(tab, date_from, date_to, group_by):
    fn = {
        'revenue':   _data_revenue,
        'orders':    _data_orders,
        'products':  _data_products,
        'customers': _data_customers,
        'inventory': _data_inventory,
        'reviews':   _data_reviews,
    }.get(tab)
    if not fn:
        raise ValueError(f'Tab không hợp lệ: {tab}')
    return fn(date_from, date_to, group_by)


def _data_revenue(df, dt, gb):
    label_expr = _date_fmt(gb)
    group_expr = _group_expr(gb)
    rows = db.session.execute(db.text(f"""
        SELECT {label_expr}           AS label,
               COUNT(*)               AS order_count,
               COALESCE(SUM(final_price),0)   AS revenue,
               COALESCE(SUM(discount),0)       AS discount,
               COALESCE(SUM(shipping_fee),0)   AS shipping_fee,
               COALESCE(SUM(final_price - discount),0) AS net_revenue
        FROM orders
        WHERE status='delivered'
          AND DATE(created_at) BETWEEN :df AND :dt
        GROUP BY {group_expr}, label
        ORDER BY {group_expr}
    """), {'df': df, 'dt': dt}).fetchall()

    kpi = db.session.execute(db.text("""
        SELECT COUNT(*)                        AS order_count,
               COUNT(DISTINCT user_id)         AS customers,
               COALESCE(SUM(final_price),0)    AS total_revenue,
               COALESCE(SUM(discount),0)       AS total_discount,
               COALESCE(SUM(shipping_fee),0)   AS total_shipping,
               COALESCE(AVG(final_price),0)    AS avg_order_value,
               COALESCE(SUM(final_price-discount),0) AS net_revenue
        FROM orders
        WHERE status='delivered'
          AND DATE(created_at) BETWEEN :df AND :dt
    """), {'df': df, 'dt': dt}).fetchone()

    cancelled = db.session.execute(db.text(
        "SELECT COUNT(*) FROM orders WHERE status='cancelled' AND DATE(created_at) BETWEEN :df AND :dt"
    ), {'df': df, 'dt': dt}).scalar()

    return {
        'kpi': {
            'order_count':    kpi[0], 'customers': kpi[1],
            'total_revenue':  int(kpi[2]), 'total_discount': int(kpi[3]),
            'total_shipping': int(kpi[4]), 'avg_order_value': int(kpi[5]),
            'net_revenue':    int(kpi[6]), 'cancelled_count': cancelled,
        },
        'chart': [{'label': r[0], 'Doanh thu': int(r[2])} for r in rows],
        'rows': [{
            'label': r[0], 'order_count': r[1],
            'revenue': int(r[2]), 'discount': int(r[3]),
            'shipping_fee': int(r[4]), 'net_revenue': int(r[5]),
        } for r in rows],
    }


def _data_orders(df, dt, gb):
    label_expr = _date_fmt(gb)
    group_expr = _group_expr(gb)
    rows = db.session.execute(db.text(f"""
        SELECT {label_expr} AS label, {group_expr} AS grp,
               COUNT(*) AS total,
               SUM(status='pending')    AS pending,
               SUM(status='processing') AS processing,
               SUM(status='shipping')   AS shipping,
               SUM(status='delivered')  AS delivered,
               SUM(status='cancelled')  AS cancelled
        FROM orders
        WHERE DATE(created_at) BETWEEN :df AND :dt
        GROUP BY grp, label ORDER BY grp
    """), {'df': df, 'dt': dt}).fetchall()

    by_status = db.session.execute(db.text("""
        SELECT status, COUNT(*) FROM orders
        WHERE DATE(created_at) BETWEEN :df AND :dt
        GROUP BY status
    """), {'df': df, 'dt': dt}).fetchall()

    status_map = {r[0]: r[1] for r in by_status}

    return {
        'by_status': status_map,
        'chart': [{
            'label': r[0], 'Tổng đơn': r[2],
            'Đã giao': r[6], 'Đã huỷ': r[7],
        } for r in rows],
        'rows': [{
            'label': r[0], 'total': r[2],
            'pending': r[3], 'processing': r[4],
            'shipping': r[5], 'delivered': r[6], 'cancelled': r[7],
            'rate': round((r[6] / r[2] * 100) if r[2] > 0 else 0, 1),
        } for r in rows],
    }


def _data_products(df, dt, gb):
    total_rev = db.session.execute(db.text("""
        SELECT COALESCE(SUM(oi.subtotal),0)
        FROM order_items oi JOIN orders o ON oi.order_id=o.id
        WHERE o.status='delivered' AND DATE(o.created_at) BETWEEN :df AND :dt
    """), {'df': df, 'dt': dt}).scalar() or 1

    rows = db.session.execute(db.text("""
        SELECT p.id, p.name,
               c.name AS category,
               COALESCE(SUM(oi.quantity),0) AS qty,
               COALESCE(SUM(oi.subtotal),0) AS revenue,
               p.quantity AS stock,
               COALESCE(p.avg_rating,0)     AS avg_rating
        FROM products p
        JOIN categories c ON p.category_id=c.id
        LEFT JOIN order_items oi ON oi.product_id=p.id
        LEFT JOIN orders o ON oi.order_id=o.id
                          AND o.status='delivered'
                          AND DATE(o.created_at) BETWEEN :df AND :dt
        WHERE p.status=1
        GROUP BY p.id, p.name, c.name, p.quantity, p.avg_rating
        HAVING qty > 0
        ORDER BY revenue DESC
    """), {'df': df, 'dt': dt}).fetchall()

    total_qty = sum(r[3] for r in rows)
    avg_rating = db.session.execute(db.text(
        "SELECT COALESCE(AVG(avg_rating),0) FROM products WHERE status=1"
    )).scalar()
    review_count = db.session.execute(db.text(
        "SELECT COUNT(*) FROM reviews WHERE DATE(created_at) BETWEEN :df AND :dt"
    ), {'df': df, 'dt': dt}).scalar()

    return {
        'kpi': {
            'sold_products': len(rows), 'total_qty': int(total_qty),
            'revenue': int(total_rev), 'avg_rating': round(float(avg_rating), 2),
            'review_count': review_count,
        },
        'rows': [{
            'id': r[0], 'name': r[1], 'category': r[2],
            'qty': r[3], 'revenue': int(r[4]),
            'pct': round(r[4] / total_rev * 100, 1),
            'stock': r[5], 'avg_rating': round(float(r[6]), 1),
        } for r in rows],
    }


def _data_customers(df, dt, gb):
    label_expr = _date_fmt(gb)
    group_expr = _group_expr(gb)

    new_rows = db.session.execute(db.text(f"""
        SELECT {label_expr} AS label, {group_expr} AS grp, COUNT(*) AS new_customers
        FROM users WHERE role='customer' AND DATE(created_at) BETWEEN :df AND :dt
        GROUP BY grp, label ORDER BY grp
    """), {'df': df, 'dt': dt}).fetchall()

    order_rows = db.session.execute(db.text(f"""
        SELECT {label_expr.replace('created_at', 'o.created_at')} AS label,
               {group_expr.replace('created_at', 'o.created_at')} AS grp,
               COUNT(DISTINCT o.user_id) AS active_customers,
               COUNT(o.id) AS orders,
               COALESCE(SUM(o.final_price),0) AS revenue
        FROM orders o WHERE o.status='delivered' AND DATE(o.created_at) BETWEEN :df AND :dt
        GROUP BY grp, label ORDER BY grp
    """), {'df': df, 'dt': dt}).fetchall()

    order_map = {r[0]: r for r in order_rows}
    rows = []
    for r in new_rows:
        om = order_map.get(r[0])
        active = om[2] if om else 0
        orders = om[3] if om else 0
        rev    = int(om[4]) if om else 0
        rows.append({
            'label': r[0], 'new_customers': r[2],
            'active_customers': active, 'orders': orders,
            'revenue': rev, 'avg_spend': int(rev / active) if active else 0,
        })

    kpi = db.session.execute(db.text("""
        SELECT COUNT(*) FROM users WHERE role='customer' AND DATE(created_at) BETWEEN :df AND :dt
    """), {'df': df, 'dt': dt}).scalar()

    active_kpi = db.session.execute(db.text("""
        SELECT COUNT(DISTINCT user_id), COALESCE(AVG(final_price),0), COUNT(*)
        FROM orders WHERE status='delivered' AND DATE(created_at) BETWEEN :df AND :dt
    """), {'df': df, 'dt': dt}).fetchone()

    avg_orders = db.session.execute(db.text("""
        SELECT AVG(cnt) FROM (
          SELECT user_id, COUNT(*) AS cnt FROM orders
          WHERE status='delivered' AND DATE(created_at) BETWEEN :df AND :dt
          GROUP BY user_id
        ) t
    """), {'df': df, 'dt': dt}).scalar() or 0

    top = db.session.execute(db.text("""
        SELECT u.id, u.name, COUNT(o.id) AS order_count, COALESCE(SUM(o.final_price),0) AS total_spend
        FROM users u JOIN orders o ON o.user_id=u.id
        WHERE o.status='delivered' AND DATE(o.created_at) BETWEEN :df AND :dt
        GROUP BY u.id, u.name ORDER BY total_spend DESC LIMIT 10
    """), {'df': df, 'dt': dt}).fetchall()

    return {
        'kpi': {
            'new_customers': kpi, 'active_customers': active_kpi[0],
            'avg_spend': int(active_kpi[1]), 'avg_orders': round(float(avg_orders), 1),
        },
        'chart': [{'label': r['label'], 'KH mới': r['new_customers']} for r in rows],
        'rows': rows,
        'top_customers': [{'id': r[0], 'name': r[1], 'order_count': r[2], 'total_spend': int(r[3])} for r in top],
    }


def _data_inventory(df, dt, gb):
    rows = db.session.execute(db.text("""
        SELECT p.id, p.name,
               c.name AS category, b.name AS brand,
               p.quantity, p.sold_count,
               COALESCE(p.sale_price, p.price) AS price,
               p.quantity * COALESCE(p.sale_price, p.price) AS stock_value
        FROM products p
        JOIN categories c ON p.category_id=c.id
        JOIN brands b     ON p.brand_id=b.id
        WHERE p.status=1
        ORDER BY p.quantity ASC
    """)).fetchall()

    total_prods  = len(rows)
    low_stock    = sum(1 for r in rows if 0 < r[4] <= 5)
    out_of_stock = sum(1 for r in rows if r[4] == 0)
    healthy      = total_prods - low_stock - out_of_stock
    stock_value  = sum(int(r[7]) for r in rows)

    return {
        'kpi': {
            'total_products': total_prods, 'low_stock': low_stock + out_of_stock,
            'healthy_stock': healthy, 'stock_value': stock_value,
        },
        'rows': [{
            'id': r[0], 'name': r[1], 'category': r[2], 'brand': r[3],
            'quantity': r[4], 'sold_count': r[5],
            'price': int(r[6]), 'stock_value': int(r[7]),
        } for r in rows],
    }


def _data_reviews(df, dt, gb):
    label_expr = _date_fmt(gb)
    group_expr = _group_expr(gb)

    chart_rows = db.session.execute(db.text(f"""
        SELECT {label_expr} AS label, {group_expr} AS grp, COUNT(*) AS cnt
        FROM reviews WHERE DATE(created_at) BETWEEN :df AND :dt
        GROUP BY grp, label ORDER BY grp
    """), {'df': df, 'dt': dt}).fetchall()

    by_star = db.session.execute(db.text("""
        SELECT rating, COUNT(*) FROM reviews
        WHERE DATE(created_at) BETWEEN :df AND :dt GROUP BY rating
    """), {'df': df, 'dt': dt}).fetchall()
    star_map = {r[0]: r[1] for r in by_star}
    total = sum(star_map.values()) or 1

    kpi_row = db.session.execute(db.text("""
        SELECT COALESCE(AVG(rating),0), COUNT(*) FROM reviews
        WHERE DATE(created_at) BETWEEN :df AND :dt
    """), {'df': df, 'dt': dt}).fetchone()

    good = star_map.get(5, 0) + star_map.get(4, 0)
    ok   = star_map.get(3, 0)
    bad  = star_map.get(2, 0) + star_map.get(1, 0)

    prod_rows = db.session.execute(db.text("""
        SELECT p.id, p.name,
               COUNT(r.id) AS total,
               ROUND(AVG(r.rating),1) AS avg,
               SUM(r.rating=5) AS s5, SUM(r.rating=4) AS s4,
               SUM(r.rating=3) AS s3, SUM(r.rating=2) AS s2,
               SUM(r.rating=1) AS s1
        FROM reviews r JOIN products p ON r.product_id=p.id
        WHERE DATE(r.created_at) BETWEEN :df AND :dt
        GROUP BY p.id, p.name ORDER BY total DESC
    """), {'df': df, 'dt': dt}).fetchall()

    return {
        'kpi': {
            'avg_rating':   round(float(kpi_row[0]), 2),
            'total_reviews': kpi_row[1],
            'good_reviews': good, 'good_pct': round(good/total*100),
            'ok_reviews':   ok,   'ok_pct':   round(ok/total*100),
            'bad_reviews':  bad,  'bad_pct':  round(bad/total*100),
        },
        'by_star': star_map,
        'chart':   [{'label': r[0], 'Đánh giá': r[2]} for r in chart_rows],
        'rows': [{
            'id': r[0], 'product_name': r[1], 'total': r[2],
            'avg': float(r[3]),
            'star_5': r[4], 'star_4': r[5], 'star_3': r[6],
            'star_2': r[7], 'star_1': r[8],
        } for r in prod_rows],
    }


# ════════════════════════════════════════════════════════════════
# Excel builders
# ════════════════════════════════════════════════════════════════
def _build_excel(tab, df, dt, gb, data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    builders = {
        'revenue':   _excel_revenue,
        'orders':    _excel_orders,
        'products':  _excel_products,
        'customers': _excel_customers,
        'inventory': _excel_inventory,
        'reviews':   _excel_reviews,
    }
    builders[tab](wb, df, dt, gb, data)
    return wb


def _excel_revenue(wb, df, dt, gb, d):
    ws = wb.create_sheet('Doanh Thu')
    cols = ['Thời gian', 'Số đơn giao', 'Doanh thu (đ)', 'Giảm giá (đ)', 'Phí ship (đ)', 'Doanh thu thuần (đ)']
    row = _write_header(ws, '📊 BÁO CÁO DOANH THU', df, dt, cols)
    _col_widths(ws, [18, 14, 20, 18, 16, 22])
    _autofreeze(ws)

    fmts = [
        (None, None),
        (FMT_VND, '374151'),
        (FMT_VND, C_TITLE),
        (FMT_VND, C_RED),
        (FMT_VND, '6B7280'),
        (FMT_VND, C_GREEN),
    ]
    rows_data = [
        (r['label'], r['order_count'], r['revenue'],
         r['discount'], r['shipping_fee'], r['net_revenue'])
        for r in d['rows']
    ]
    _write_rows(ws, row, rows_data, fmts)

    k = d['kpi']
    total_row = row + len(rows_data)
    _write_total(ws, total_row, [
        'TỔNG CỘNG', k['order_count'], k['total_revenue'],
        k['total_discount'], k['total_shipping'], k['net_revenue'],
    ])
    for ci, fmt in enumerate([None, FMT_VND, FMT_VND, FMT_VND, FMT_VND, FMT_VND], 1):
        if fmt: ws.cell(total_row, ci).number_format = fmt

    # KPI summary block
    _excel_kpi_block(ws, total_row + 2, [
        ('Tổng doanh thu', k['total_revenue'], FMT_VND),
        ('Giá trị TB/đơn', k['avg_order_value'], FMT_VND),
        ('Tổng giảm giá',  k['total_discount'], FMT_VND),
        ('Đơn bị huỷ',     k['cancelled_count'], None),
    ], len(cols))


def _excel_orders(wb, df, dt, gb, d):
    ws = wb.create_sheet('Đơn Hàng')
    cols = ['Thời gian', 'Tổng đơn', 'Chờ xử lý', 'Đang xử lý', 'Đang giao', 'Đã giao', 'Đã huỷ', 'Tỷ lệ giao (%)']
    row = _write_header(ws, '📊 BÁO CÁO ĐƠN HÀNG', df, dt, cols)
    _col_widths(ws, [18, 12, 14, 14, 14, 12, 12, 18])
    _autofreeze(ws)

    fmts = [
        (None, None), (None, None),
        (None, C_AMBER), (None, C_TITLE), (None, '8B5CF6'),
        (None, C_GREEN), (None, C_RED), (FMT_PCT, None),
    ]
    rows_data = [
        (r['label'], r['total'], r['pending'], r['processing'],
         r['shipping'], r['delivered'], r['cancelled'], r['rate'] / 100)
        for r in d['rows']
    ]
    _write_rows(ws, row, rows_data, fmts)

    # Status summary
    ws2 = wb.create_sheet('Phân Bổ Trạng Thái')
    ws2['A1'] = 'Trạng thái'; ws2['B1'] = 'Số đơn'; ws2['C1'] = 'Tỷ lệ'
    for ci, h in enumerate(['A1', 'B1', 'C1']):
        ws2[h].font = _hdr_font(); ws2[h].fill = _fill(C_HEADER); ws2[h].alignment = _center()
    labels = {'pending': 'Chờ xử lý', 'processing': 'Đang xử lý', 'shipping': 'Đang giao', 'delivered': 'Đã giao', 'cancelled': 'Đã huỷ'}
    total_ord = sum(d['by_status'].values()) or 1
    for ri, (k, l) in enumerate(labels.items(), 2):
        v = d['by_status'].get(k, 0)
        ws2.cell(ri, 1, l).font = _body_font()
        ws2.cell(ri, 2, v).font = _body_font(); ws2.cell(ri, 2).alignment = _right()
        ws2.cell(ri, 3, v / total_ord).number_format = FMT_PCT; ws2.cell(ri, 3).alignment = _right()
    _col_widths(ws2, [18, 12, 14])


def _excel_products(wb, df, dt, gb, d):
    ws = wb.create_sheet('Sản Phẩm')
    cols = ['#', 'Tên sản phẩm', 'Danh mục', 'SL bán', 'Doanh thu (đ)', '% Tổng DT', 'Tồn kho', 'Rating']
    row = _write_header(ws, '📊 BÁO CÁO SẢN PHẨM BÁN CHẠY', df, dt, cols)
    _col_widths(ws, [5, 40, 18, 10, 22, 12, 10, 10])
    _autofreeze(ws)

    fmts = [
        (None, '9CA3AF'), (None, None), (None, '6B7280'),
        (None, None), (FMT_VND, C_TITLE), (FMT_PCT, '6B7280'),
        (None, None), (None, C_AMBER),
    ]
    rows_data = [
        (i + 1, r['name'], r['category'], r['qty'], r['revenue'],
         r['pct'] / 100, r['stock'], r['avg_rating'])
        for i, r in enumerate(d['rows'])
    ]
    _write_rows(ws, row, rows_data, fmts)
    # Color stock cells
    for ri, r in enumerate(d['rows']):
        cell = ws.cell(row + ri, 7)
        cell.font = _body_font(bold=True, color=C_RED if r['stock'] <= 5 else C_GREEN)


def _excel_customers(wb, df, dt, gb, d):
    ws = wb.create_sheet('Khách Hàng')
    cols = ['Thời gian', 'KH mới', 'KH có đơn', 'Tổng đơn', 'Doanh thu (đ)', 'Chi tiêu TB (đ)']
    row = _write_header(ws, '📊 BÁO CÁO KHÁCH HÀNG', df, dt, cols)
    _col_widths(ws, [18, 12, 14, 12, 22, 20])
    _autofreeze(ws)

    fmts = [
        (None, None), (None, C_GREEN), (None, C_TITLE),
        (None, None), (FMT_VND, C_TITLE), (FMT_VND, '6B7280'),
    ]
    rows_data = [
        (r['label'], r['new_customers'], r['active_customers'],
         r['orders'], r['revenue'], r['avg_spend'])
        for r in d['rows']
    ]
    _write_rows(ws, row, rows_data, fmts)

    # Top customers sheet
    ws2 = wb.create_sheet('Top Khách Hàng')
    headers = ['#', 'Họ tên', 'Số đơn', 'Tổng chi tiêu (đ)']
    for ci, h in enumerate(headers, 1):
        c = ws2.cell(1, ci, h)
        c.font = _hdr_font(); c.fill = _fill(C_HEADER); c.alignment = _center()
    for ri, c in enumerate(d.get('top_customers', []), 2):
        ws2.cell(ri, 1, ri - 1)
        ws2.cell(ri, 2, c['name'])
        ws2.cell(ri, 3, c['order_count']).alignment = _right()
        cell = ws2.cell(ri, 4, c['total_spend'])
        cell.number_format = FMT_VND; cell.alignment = _right()
        cell.font = _body_font(bold=True, color=C_TITLE)
    _col_widths(ws2, [5, 30, 12, 24])


def _excel_inventory(wb, df, dt, gb, d):
    ws = wb.create_sheet('Tồn Kho')
    cols = ['#', 'Sản phẩm', 'Danh mục', 'Thương hiệu', 'Tồn kho', 'Đã bán', 'Giá bán (đ)', 'Giá trị tồn (đ)', 'Trạng thái']
    row = _write_header(ws, '📊 BÁO CÁO TỒN KHO', df, dt, cols)
    _col_widths(ws, [5, 38, 18, 14, 10, 10, 20, 22, 14])
    _autofreeze(ws)

    for ri, r in enumerate(d['rows']):
        fill = _fill(C_ODD) if ri % 2 == 0 else None
        vals = [ri + 1, r['name'], r['category'], r['brand'],
                r['quantity'], r['sold_count'], r['price'], r['stock_value'],
                'Hết hàng' if r['quantity'] == 0 else ('Sắp hết' if r['quantity'] <= 5 else 'Còn hàng')]
        fmts_row = [None, None, None, None, None, None, FMT_VND, FMT_VND, None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts_row), 1):
            cell = ws.cell(row + ri, ci, val)
            cell.font   = _body_font()
            cell.border = _border()
            if fill: cell.fill = fill
            if fmt:  cell.number_format = fmt
            if ci > 1: cell.alignment = _right()
            else:       cell.alignment = Alignment(horizontal='left', vertical='center')
        # Color stock quantity
        q_cell = ws.cell(row + ri, 5)
        q_cell.font = _body_font(bold=True, color=C_RED if r['quantity'] <= 5 else C_GREEN)
        # Status color
        s_cell = ws.cell(row + ri, 9)
        s_cell.font = _body_font(bold=True, color=C_RED if r['quantity'] == 0 else (C_AMBER if r['quantity'] <= 5 else C_GREEN))


def _excel_reviews(wb, df, dt, gb, d):
    ws = wb.create_sheet('Đánh Giá SP')
    cols = ['#', 'Sản phẩm', 'SL đánh giá', 'Rating TB', '5 sao', '4 sao', '3 sao', '2 sao', '1 sao']
    row = _write_header(ws, '📊 BÁO CÁO ĐÁNH GIÁ SẢN PHẨM', df, dt, cols)
    _col_widths(ws, [5, 40, 14, 12, 10, 10, 10, 10, 10])
    _autofreeze(ws)

    for ri, r in enumerate(d['rows']):
        fill = _fill(C_ODD) if ri % 2 == 0 else None
        vals = [ri + 1, r['product_name'], r['total'], r['avg'],
                r['star_5'], r['star_4'], r['star_3'], r['star_2'], r['star_1']]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row + ri, ci, val)
            cell.font = _body_font(); cell.border = _border()
            if fill: cell.fill = fill
            if ci > 1: cell.alignment = _right()
            else:       cell.alignment = Alignment(horizontal='left', vertical='center')
        # Rating color
        avg_cell = ws.cell(row + ri, 4)
        avg = r['avg']
        avg_cell.font = _body_font(bold=True, color=C_GREEN if avg >= 4 else (C_AMBER if avg >= 3 else C_RED))

    # Star distribution sheet
    ws2 = wb.create_sheet('Phân Bổ Sao')
    for ci, h in enumerate(['Số sao', 'Số đánh giá', 'Tỷ lệ'], 1):
        c = ws2.cell(1, ci, h)
        c.font = _hdr_font(); c.fill = _fill(C_HEADER); c.alignment = _center()
    total = d['kpi']['total_reviews'] or 1
    for ri, star in enumerate([5, 4, 3, 2, 1], 2):
        v = d['by_star'].get(star, 0)
        ws2.cell(ri, 1, f'{star} sao').alignment = _center()
        ws2.cell(ri, 2, v).alignment = _right()
        pct_cell = ws2.cell(ri, 3, v / total)
        pct_cell.number_format = FMT_PCT; pct_cell.alignment = _right()
    _col_widths(ws2, [12, 16, 14])


def _excel_kpi_block(ws, start_row, items, n_cols):
    """Draw a KPI summary block below the main table."""
    ws.cell(start_row, 1, 'TÓM TẮT KPI').font = Font(name='Arial', bold=True, size=11, color=C_TITLE)
    for i, (label, value, fmt) in enumerate(items, 1):
        r = start_row + i
        ws.cell(r, 1, label).font = _body_font(bold=True)
        cell = ws.cell(r, 2, value)
        cell.font = _body_font(bold=True, color=C_TITLE)
        if fmt: cell.number_format = fmt
        cell.alignment = _right()